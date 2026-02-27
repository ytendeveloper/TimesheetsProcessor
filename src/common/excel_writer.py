"""
Excel workbook writer for Timesheets Processor.

Manages per-employee Excel (.xlsx) files with:
- One tab per year, created lazily when needed
- Weekly mode: rows prepopulated with Sun-Sat date ranges for the year
- Monthly mode: rows prepopulated with Jan-Dec month ranges
- Update specific rows by matching dates

Uses openpyxl for all Excel operations.
"""

import calendar
import logging
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger("timesheets_processor")

# Header styling constants
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Date range generation
# ---------------------------------------------------------------------------

def get_weekly_ranges(year: int) -> list[tuple[date, date]]:
    """
    Generate all Sun-Sat week ranges that overlap with the given year.

    Includes partial weeks at year boundaries so no days are missed.

    Args:
        year: The calendar year.

    Returns:
        List of (sunday_start, saturday_end) tuples.
    """
    ranges = []

    # Find the first Sunday on or before Jan 1
    jan1 = date(year, 1, 1)
    days_since_sunday = (jan1.weekday() + 1) % 7
    week_start = jan1 - timedelta(days=days_since_sunday)

    while True:
        week_end = week_start + timedelta(days=6)
        # Include this week if any part of it falls within the target year
        if week_end.year < year:
            week_start += timedelta(days=7)
            continue
        if week_start.year > year:
            break
        ranges.append((week_start, week_end))
        week_start += timedelta(days=7)

    return ranges


def get_monthly_ranges(year: int) -> list[tuple[date, date]]:
    """
    Generate first-to-last-day ranges for each month of the year.

    Args:
        year: The calendar year.

    Returns:
        List of 12 (month_start, month_end) tuples.
    """
    ranges = []
    for month in range(1, 13):
        start = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end = date(year, month, last_day)
        ranges.append((start, end))
    return ranges


# ---------------------------------------------------------------------------
# Workbook management
# ---------------------------------------------------------------------------

def get_or_create_workbook(filepath: Path) -> Workbook:
    """
    Load an existing Excel workbook or create a new one.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        openpyxl Workbook instance.
    """
    if filepath.exists():
        logger.debug("Loading existing workbook: %s", filepath)
        return load_workbook(filepath)

    logger.debug("Creating new workbook: %s", filepath)
    return Workbook()


def get_or_create_year_sheet(wb: Workbook, year: int,
                             mode: str = "weekly"):
    """
    Get an existing year tab or create and prepopulate it.

    Args:
        wb: The openpyxl Workbook.
        year: The year for this tab (e.g. 2025).
        mode: "weekly" for Sun-Sat rows, "monthly" for Jan-Dec rows.

    Returns:
        The openpyxl Worksheet for this year.
    """
    sheet_name = str(year)

    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    # Remove the default "Sheet" if it's the only one and is empty
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        ws = wb["Sheet"]
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
            del wb["Sheet"]

    ws = wb.create_sheet(title=sheet_name)

    if mode == "weekly":
        _populate_weekly_sheet(ws, year)
    else:
        _populate_monthly_sheet(ws, year)

    logger.info("Created %s tab for year %d with %s date ranges", sheet_name, year, mode)
    return ws


def _style_header_row(ws, num_cols: int):
    """Apply styling to the header row (row 1)."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _populate_weekly_sheet(ws, year: int):
    """
    Prepopulate a worksheet with weekly Sun-Sat date ranges.

    Columns: Week Start (Sun) | Week End (Sat) | Hours Worked | Project/Client | Notes
    """
    headers = ["Week Start (Sun)", "Week End (Sat)", "Hours Worked", "Project/Client", "Notes"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for week_start, week_end in get_weekly_ranges(year):
        ws.append([week_start, week_end, None, None, None])

    # Format date columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = "MM/DD/YYYY"

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30


def _populate_monthly_sheet(ws, year: int):
    """
    Prepopulate a worksheet with monthly date ranges.

    Columns: Month | Start Date | End Date | Hours Worked | Project/Client | Notes
    """
    headers = ["Month", "Start Date", "End Date", "Hours Worked", "Project/Client", "Notes"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    for i, (start, end) in enumerate(get_monthly_ranges(year)):
        ws.append([month_names[i], start, end, None, None, None])

    # Format date columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "MM/DD/YYYY"

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 30


# ---------------------------------------------------------------------------
# Row updates
# ---------------------------------------------------------------------------

def update_weekly_row(ws, week_start_date: date, hours: float,
                      project: Optional[str]) -> bool:
    """
    Find the row matching a week start date and update hours/project.

    Column A = Week Start (Sun), Column C = Hours Worked, Column D = Project/Client.

    Args:
        ws: The year worksheet.
        week_start_date: The Sunday start date to match.
        hours: Hours worked to write.
        project: Project/client name to write.

    Returns:
        True if a matching row was found and updated, False otherwise.
    """
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell_val = row[0].value   # Column A
        # openpyxl may return datetime instead of date
        if hasattr(cell_val, "date"):
            cell_val = cell_val.date()
        if cell_val == week_start_date:
            row[2].value = hours      # Column C: Hours Worked
            if project:
                row[3].value = project    # Column D: Project/Client
            logger.debug(
                "Updated weekly row: %s → %.1f hours, project=%s",
                week_start_date, hours, project
            )
            return True
    return False


def update_monthly_row(ws, year: int, month: int, hours: float,
                       project: Optional[str]) -> bool:
    """
    Find the row for a given month and update hours/project.

    Column B = Start Date (used to identify month), Column D = Hours, Column E = Project.

    Args:
        ws: The year worksheet.
        year: The year.
        month: The month (1-12).
        hours: Hours worked to write.
        project: Project/client name to write.

    Returns:
        True if a matching row was found and updated, False otherwise.
    """
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell_val = row[1].value   # Column B: Start Date
        if hasattr(cell_val, "date"):
            cell_val = cell_val.date()
        if cell_val and cell_val.month == month and cell_val.year == year:
            row[3].value = hours      # Column D: Hours Worked
            if project:
                row[4].value = project    # Column E: Project/Client
            logger.debug(
                "Updated monthly row: %d/%d → %.1f hours, project=%s",
                month, year, hours, project
            )
            return True
    return False


def save_workbook(wb: Workbook, filepath: Path):
    """
    Sort tabs chronologically and save the workbook.

    Args:
        wb: The openpyxl Workbook.
        filepath: Path to save the .xlsx file.
    """
    # Sort sheets by name (year strings sort correctly)
    wb._sheets.sort(key=lambda ws: ws.title)
    wb.save(filepath)
    logger.info("Saved workbook: %s", filepath)
