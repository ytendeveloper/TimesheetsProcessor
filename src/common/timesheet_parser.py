"""
Timesheet file parser for Timesheets Processor.

Extracts text from PDF and image (PNG/JPG) timesheet files, then parses
structured data (dates, hours, project) using regex patterns.

Supports multi-entry timesheets where a single file contains multiple
weekly time card entries in the format:
    DD-Mon-YYYY - DD-Mon-YYYY
    Total Hours:XX.XX
    Approved/Submitted

Supported file types:
- PDF: text extracted via pdfplumber (tables + raw text)
- PNG/JPG/JPEG: text extracted via pytesseract OCR with OpenCV preprocessing
"""

import calendar
import io
import logging
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional, Tuple

import cv2
import fitz  # PyMuPDF — renders PDF pages to images for OCR fallback
import numpy as np
import pdfplumber
import pytesseract
from PIL import Image

logger = logging.getLogger("timesheets_processor")

SUPPORTED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".xlsx", ".zip"}


@dataclass
class TimesheetEntry:
    """A single timesheet entry (one week or one month of hours)."""
    start_date: date
    end_date: date
    hours: float
    period_type: str            # "weekly" or "monthly"
    status: Optional[str]       # "Approved", "Submitted", etc.


@dataclass
class TimesheetFileData:
    """All data extracted from a single timesheet file."""
    employee_name: Optional[str]
    project: Optional[str]
    entries: List[TimesheetEntry]
    raw_text: str


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

def extract_text(file_path: Path) -> str:
    """
    Extract text from a timesheet file, routing by file extension.

    Args:
        file_path: Path to a PDF or image file.

    Returns:
        Extracted text string.

    Raises:
        ValueError: If the file extension is not supported.
    """
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return _extract_text_from_pdf(file_path)
    elif ext in {".png", ".jpg", ".jpeg"}:
        return _extract_text_from_image(file_path)
    elif ext == ".xlsx":
        return ""  # xlsx files are parsed directly, not via text extraction
    elif ext == ".zip":
        return ""  # zip files are handled by extract_from_zip()
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _extract_text_from_pdf(file_path: Path) -> str:
    """
    Extract text from a PDF using pdfplumber.

    Uses raw text extraction (not table extraction) to preserve the
    line-by-line structure needed for parsing date/hours blocks.

    Args:
        file_path: Path to the PDF file.

    Returns:
        Concatenated text from all pages.
    """
    all_text = []

    with pdfplumber.open(str(file_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if text:
                all_text.append(text)

    result = "\n".join(all_text)

    # OCR fallback: if pdfplumber extracted no text, render pages as images
    if not result.strip():
        logger.debug("No text from pdfplumber for %s — trying OCR fallback", file_path.name)
        result = _ocr_pdf_pages(file_path)

    logger.debug("PDF extraction from %s: %d characters", file_path.name, len(result))
    return result


def _preprocess_for_ocr(img: np.ndarray) -> Image.Image:
    """
    Preprocess an OpenCV image for OCR: upscale 2x, grayscale, adaptive threshold.

    The 2x upscale improves tesseract accuracy on screenshot-quality images,
    and adaptive thresholding handles colored UI elements better than Otsu.
    """
    h, w = img.shape[:2]
    img = cv2.resize(img, (w * 2, h * 2), interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 10
    )
    return Image.fromarray(thresh)


def _ocr_pdf_pages(file_path: Path) -> str:
    """
    Render each PDF page to an image at 300 DPI and run OCR via tesseract.

    Used as a fallback when pdfplumber cannot extract text (image-based PDFs).
    """
    if not shutil.which("tesseract"):
        raise RuntimeError(
            "Tesseract OCR is not installed. Install with: brew install tesseract"
        )

    all_text = []
    doc = fitz.open(str(file_path))
    for page_num, page in enumerate(doc):
        # Render at 300 DPI (default is 72; scale factor = 300/72 ≈ 4.17)
        mat = fitz.Matrix(300 / 72, 300 / 72)
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes("png")

        pil_img = Image.open(io.BytesIO(img_data))
        cv_img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
        ocr_img = _preprocess_for_ocr(cv_img)

        text = pytesseract.image_to_string(ocr_img, config="--oem 3 --psm 6")
        if text.strip():
            all_text.append(text)
        logger.debug("OCR fallback page %d of %s: %d chars", page_num + 1, file_path.name, len(text))

    doc.close()
    return "\n".join(all_text)


def _extract_text_from_image(file_path: Path) -> str:
    """
    Extract text from an image using pytesseract OCR with OpenCV preprocessing.

    Preprocessing: 2x upscale -> grayscale -> adaptive threshold binarization.

    Args:
        file_path: Path to the image file.

    Returns:
        Extracted text string.
    """
    if not shutil.which("tesseract"):
        raise RuntimeError(
            "Tesseract OCR is not installed. Install with: brew install tesseract"
        )

    img = cv2.imread(str(file_path))
    if img is None:
        raise ValueError(f"Could not read image file: {file_path}")

    # Skip tiny images (email signature logos, social media icons)
    h, w = img.shape[:2]
    if w < 300 or h < 300:
        logger.debug("Skipping small image %s (%dx%d) — likely not a timesheet", file_path.name, w, h)
        return ""

    ocr_img = _preprocess_for_ocr(img)
    text = pytesseract.image_to_string(ocr_img, config="--oem 3 --psm 6")

    logger.debug("OCR extraction from %s: %d characters", file_path.name, len(text))
    return text


def parse_xlsx_timesheet(file_path: Path) -> TimesheetFileData:
    """
    Parse a Techno-Comp style Excel timesheet directly from cell values.

    Expected layout:
    - Row 7: CONSULTANT NAME (col 6), Date (col 15)
    - Row 13: PERIOD START DATE (col 5), PERIOD END DATE (col 9)
    - Row 18: Week 1 billable hours (cols 3,5,7,9,11,13,15 = Mon-Sun)
    - Row 24: Week 2 billable hours (same columns)
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb.active

    # Extract employee name
    employee_name = None
    name_val = ws.cell(row=7, column=6).value
    if name_val and str(name_val).strip():
        employee_name = str(name_val).strip()

    # Extract period dates
    start_val = ws.cell(row=13, column=5).value
    end_val = ws.cell(row=13, column=9).value
    if not start_val or not end_val:
        logger.debug("No Techno-Comp period dates in %s — trying daily-log format", file_path.name)
        result = _parse_xlsx_daily_log(wb, file_path)
        wb.close()
        return result

    # Parse dates (may be datetime objects or strings)
    if isinstance(start_val, datetime):
        period_start = start_val.date()
    elif isinstance(start_val, date):
        period_start = start_val
    else:
        period_start = _parse_date_flexible(str(start_val))

    if isinstance(end_val, datetime):
        period_end = end_val.date()
    elif isinstance(end_val, date):
        period_end = end_val
    else:
        period_end = _parse_date_flexible(str(end_val))

    if not period_start or not period_end:
        logger.warning("Could not parse period dates in %s", file_path.name)
        return TimesheetFileData(employee_name=employee_name, project=None, entries=[], raw_text="")

    # Extract weekly hours from rows 18 and 24 (7 day columns each)
    hour_cols = [3, 5, 7, 9, 11, 13, 15]
    entries = []

    for week_idx, hour_row in enumerate([18, 24]):
        total = 0.0
        for col in hour_cols:
            val = ws.cell(row=hour_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        if total > 0:
            # Week 1 starts at period_start, week 2 starts 7 days later
            week_start = period_start + timedelta(days=7 * week_idx)
            week_end = week_start + timedelta(days=6)
            entries.append(TimesheetEntry(
                start_date=week_start, end_date=week_end,
                hours=total, period_type="weekly", status=None,
            ))

    wb.close()
    logger.debug("Parsed xlsx %s: %d entries, employee: %s", file_path.name, len(entries), employee_name)

    return TimesheetFileData(
        employee_name=employee_name, project=None,
        entries=entries, raw_text=f"[xlsx: {file_path.name}]",
    )


def _parse_xlsx_daily_log(wb, file_path: Path) -> TimesheetFileData:
    """
    Parse non-Techno-Comp xlsx timesheets.  Tries two sub-formats:

    1. **Daily log** — each row has a single date (col C) and hours (col D).
       Groups by week (Sunday-start) and returns one TimesheetEntry per week.

    2. **Biweekly payroll summary** — col C contains a date range string
       like ``"01/04 - 01/17"`` and cols D/E hold Week1/Week2 hours.
       Returns one TimesheetEntry per week with hours > 0.
    """
    from collections import defaultdict

    ws = wb.active
    daily: dict[date, float] = {}
    biweekly_entries: List[TimesheetEntry] = []

    # ------------------------------------------------------------------
    # Biweekly date-range pattern: "MM/DD - MM/DD" in col C
    # ------------------------------------------------------------------
    _biweekly_range = re.compile(
        r"(\d{1,2}/\d{1,2})\s*-\s*(\d{1,2}/\d{1,2})"
    )

    for row in ws.iter_rows(min_row=2, max_col=10, values_only=False):
        date_cell = row[2].value if len(row) > 2 else None  # col C
        hours_cell = row[3].value if len(row) > 3 else None  # col D
        hours2_cell = row[4].value if len(row) > 4 else None  # col E (Week2)

        if date_cell is None:
            continue

        # --- Try biweekly date-range format first ---
        if isinstance(date_cell, str):
            rng = _biweekly_range.search(date_cell)
            if rng:
                # Infer year from sheet title or filename
                year = None
                if ws.title and re.search(r"\d{4}", ws.title):
                    year = int(re.search(r"\d{4}", ws.title).group())
                if not year:
                    yr_match = re.search(r"\d{4}", file_path.stem)
                    if yr_match:
                        year = int(yr_match.group())
                if not year:
                    year = date.today().year

                try:
                    start_str = f"{rng.group(1)}/{year}"
                    end_str = f"{rng.group(2)}/{year}"
                    week1_start = datetime.strptime(start_str, "%m/%d/%Y").date()
                    week2_start = week1_start + timedelta(days=7)
                    week1_end = week1_start + timedelta(days=6)
                    week2_end = week2_start + timedelta(days=6)
                except ValueError:
                    continue

                # Week 1 hours (col D)
                w1 = 0.0
                if hours_cell is not None:
                    try:
                        w1 = float(hours_cell)
                    except (ValueError, TypeError):
                        pass
                # Week 2 hours (col E)
                w2 = 0.0
                if hours2_cell is not None:
                    try:
                        w2 = float(hours2_cell)
                    except (ValueError, TypeError):
                        pass

                if w1 > 0:
                    biweekly_entries.append(TimesheetEntry(
                        start_date=week1_start, end_date=week1_end,
                        hours=w1, period_type="weekly", status=None,
                    ))
                if w2 > 0:
                    biweekly_entries.append(TimesheetEntry(
                        start_date=week2_start, end_date=week2_end,
                        hours=w2, period_type="weekly", status=None,
                    ))
                continue  # handled as biweekly, skip daily logic

        if hours_cell is None:
            continue

        # --- Daily log format: single date in col C ---
        if isinstance(date_cell, datetime):
            d = date_cell.date()
        elif isinstance(date_cell, date):
            d = date_cell
        else:
            d = _parse_date_flexible(str(date_cell))
        if not d:
            continue

        try:
            h = float(hours_cell)
        except (ValueError, TypeError):
            continue
        if h <= 0:
            continue

        daily[d] = daily.get(d, 0.0) + h

    # Return biweekly entries if found
    if biweekly_entries:
        logger.debug("Parsed biweekly xlsx %s: %d weekly entries",
                     file_path.name, len(biweekly_entries))
        return TimesheetFileData(
            employee_name=None, project=None,
            entries=biweekly_entries, raw_text=f"[xlsx biweekly: {file_path.name}]",
        )

    if not daily:
        logger.warning("No daily entries found in %s", file_path.name)
        return TimesheetFileData(employee_name=None, project=None, entries=[], raw_text="")

    # Group by week (Sunday-start)
    weekly: dict[date, float] = defaultdict(float)
    for d, h in daily.items():
        week_sun = normalize_to_week_start(d)
        weekly[week_sun] += h

    entries = []
    for week_start in sorted(weekly):
        entries.append(TimesheetEntry(
            start_date=week_start,
            end_date=week_start + timedelta(days=6),
            hours=round(weekly[week_start], 2),
            period_type="weekly",
            status=None,
        ))

    logger.debug("Parsed daily-log xlsx %s: %d weeks from %d daily rows",
                 file_path.name, len(entries), len(daily))
    return TimesheetFileData(
        employee_name=None, project=None,
        entries=entries, raw_text=f"[xlsx daily-log: {file_path.name}]",
    )


def extract_from_zip(zip_path: Path, dest_dir: Path) -> List[Path]:
    """
    Extract supported files from a zip archive into dest_dir.

    Skips macOS resource fork files (__MACOSX/) and temp files (~$).
    Returns list of extracted file paths.
    """
    extracted = []
    with zipfile.ZipFile(str(zip_path), "r") as zf:
        for info in zf.infolist():
            # Skip directories, macOS metadata, and temp files
            if info.is_dir():
                continue
            basename = Path(info.filename).name
            if basename.startswith("._") or basename.startswith("~$") or "__MACOSX" in info.filename:
                continue
            ext = Path(basename).suffix.lower()
            if ext not in SUPPORTED_EXTENSIONS:
                continue

            # Extract to dest_dir with a flat name (prefix from zip entry)
            dest = dest_dir / f"{zip_path.stem}_{basename}"
            with zf.open(info) as src, open(str(dest), "wb") as dst:
                dst.write(src.read())
            extracted.append(dest)
            logger.debug("Extracted %s from %s", basename, zip_path.name)

    logger.info("Extracted %d files from %s", len(extracted), zip_path.name)
    return extracted


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

def _parse_date_flexible(date_str: str) -> Optional[date]:
    """
    Try multiple date formats to parse a date string.

    Handles:
    - DD-Mon-YYYY (e.g. 05-Mar-2025) — primary format from timesheets
    - MM/DD/YYYY, MM-DD-YYYY, M/D/YY
    - Month DD, YYYY
    """
    date_str = date_str.strip().rstrip(",")
    # OCR sometimes reads "/" as "'" — normalize before parsing
    date_str = date_str.replace("'", "/")
    formats = [
        "%d-%b-%Y",     # 05-Mar-2025 (primary format in these timesheets)
        "%d-%B-%Y",     # 05-March-2025
        "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
        "%Y-%m-%d",
        "%B %d, %Y", "%b %d, %Y", "%B %d %Y", "%b %d %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Multi-entry parsing
# ---------------------------------------------------------------------------

# Pattern for date ranges like "05-Mar-2025 - 09-Mar-2025"
# Captures: DD-Mon-YYYY - DD-Mon-YYYY
_DATE_RANGE_PATTERN = re.compile(
    r"(\d{1,2}-[A-Za-z]{3,9}-\d{4})\s*[-–—]\s*(\d{1,2}-[A-Za-z]{3,9}-\d{4})"
)

# Also support numeric date ranges: MM/DD/YYYY - MM/DD/YYYY
_NUMERIC_DATE_RANGE_PATTERN = re.compile(
    r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s*(?:to|through|[-–—])\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})"
)

# Pattern for "From Sunday MM/DD/YYYY to Saturday MM/DD/YYYY" (PeopleSoft format)
# OCR sometimes drops the "F" → "rom Sunday ...", or misreads "/" as "'"
_FROM_TO_DATE_PATTERN = re.compile(
    r"F?rom\s+\w+\s+(\d{1,2}[/']\d{1,2}[/']\d{2,4})\s+to\s+\w+\s+(\d{1,2}[/']\d{1,2}[/']\d{2,4})",
    re.IGNORECASE
)

# Pattern for "Total Hours:XX.XX" (with or without spaces)
_HOURS_PATTERN = re.compile(
    r"Total\s*Hours?\s*[:\-]?\s*(\d+(?:\.\d+)?)", re.IGNORECASE
)

# Pattern for "Reported Hours XX.XXX" (PeopleSoft format)
# OCR may insert junk chars (e.g. "§") between "Hours" and the number
_REPORTED_HOURS_PATTERN = re.compile(
    r"Reported\s+Hours\s+\W*(\d+(?:\.\d+)?)", re.IGNORECASE
)

# PeopleSoft grid line: "7.500 7.500 7.500 7.500 0.000  30.000 CWR Hours Reg Hrs"
# Captures the total (30.000) from CWR rows, ignoring CWADJ adjustment rows.
_CWR_GRID_TOTAL_PATTERN = re.compile(
    r"(\d+\.\d+)\s*[|)\]}\s]*CWR\b", re.IGNORECASE
)

# Fallback: sum hours from "Reported Time Status" detail rows when the grid
# line is OCR-garbled.  Matches lines like "Approved 7.500 CWR Reg Hours ..."
_CWR_DETAIL_ROW_PATTERN = re.compile(
    r"Approved\s+(\d+\.\d+)\s+CWR\b", re.IGNORECASE
)

# TRC codes to ignore (system adjustment entries, not actual work hours)
_IGNORED_TRC_CODES = {"CWADJ"}

# Pattern for status lines
_STATUS_PATTERN = re.compile(
    r"\b(Approved|Submitted|Pending|Rejected|Draft)\b", re.IGNORECASE
)

# Employee name pattern: look for name after "AA\n" header block
# Also matches PeopleSoft format: "Name Surname Employee ID NNNNNN"
_NAME_PATTERN = re.compile(
    r"(?:AA\n|Employee\s*[:\-]\s*|Name\s*[:\-]\s*)([A-Za-z][\w\s\-']+?)(?:\n|$)"
    r"|^Timesheet\n([A-Za-z][A-Za-z\s]+?)\s+Employee\s+ID",
    re.IGNORECASE | re.MULTILINE
)

# "Totals:  ...  40.00" (Unanet EaZyTyme format)
# OCR may insert artifacts like _] |  [ etc. — just grab the last decimal on the line
_TOTALS_PATTERN = re.compile(
    r"Totals?\b.*?(\d+\.\d{2})[)\]\[|\s]*$", re.IGNORECASE | re.MULTILINE
)

# "Total XX.XX" standalone (MBO Partners format — no "Hours" keyword)
_TOTAL_ONLY_PATTERN = re.compile(
    r"\bTotal\s+(\d+\.\d{1,2})\b"
)

# "timesheet for Aug 04 - Aug 10, 2025" (MBO Partners title format)
_MBO_TITLE_DATE_PATTERN = re.compile(
    r"timesheet\s+for\s+(\w{3,9})\s+(\d{1,2})\s*[-–—]\s*(\w{3,9})\s+(\d{1,2}),?\s*(\d{4})",
    re.IGNORECASE,
)

# "10/19/2025 — 10/25/2025    40    LOCKED" (Unanet Time-List inline format)
_INLINE_HOURS_AFTER_DATES = re.compile(
    r"(\d{1,2}/\d{1,2}/\d{4})\s*[—–\-]+\s*(\d{1,2}/\d{1,2}/\d{4})\s*[|\s]+(\d+(?:\.\d+)?)\s*[|\s]+"
)

# Unanet detail grid row: "BILL |RT | 8.00 | 8.00 | ... | 40.00;"
# Captures the last decimal number on a BILL line (the row total).
_UNANET_DETAIL_GRID_PATTERN = re.compile(
    r"BILL\b.*?(\d+\.\d{2})[;|,]?\s*$", re.IGNORECASE | re.MULTILINE
)

# "Timesheet for NAME (MM/DD/YYYY - MM/DD/YYYY)" — date range in parentheses
_PAREN_DATE_RANGE = re.compile(
    r"\((\d{1,2}/\d{1,2}/\d{4})\s*[-–—]\s*(\d{1,2}/\d{1,2}/\d{4})\)"
)

# Project/client extraction pattern for PeopleSoft timesheets.
# Matches the category line after "Employee ID NNNNNN\n", e.g.:
#   "VLO Category 2B 40 Empl Record 0"  → captures "VLO Category 2B"
#   "Full Service Cat1 40 Empl Record 0" → captures "Full Service Cat1"
# The trailing number (standard hours) and "Empl Record" are stripped.
_PEOPLESOFT_CATEGORY_PATTERN = re.compile(
    r"Employee\s+ID\s+\d+\n(.+?)\s+\d+\s+Empl\s+Record",
    re.IGNORECASE
)


def parse_all_entries(text: str) -> TimesheetFileData:
    """
    Parse all timesheet entries from raw text.

    Handles multi-entry timesheets where a single file contains many
    weekly time card blocks. Each block has the format:
        DD-Mon-YYYY - DD-Mon-YYYY
        Total Hours:XX.XX
        Approved/Submitted

    Also extracts the employee name from header text.

    Args:
        text: Raw text extracted from a timesheet file.

    Returns:
        TimesheetFileData with employee_name and list of TimesheetEntry objects.
    """
    # Extract employee name from header
    employee_name = None
    name_match = _NAME_PATTERN.search(text)
    if name_match:
        # group(1) is from AA/Employee/Name patterns, group(2) from PeopleSoft
        employee_name = (name_match.group(1) or name_match.group(2) or "").strip()
        if not employee_name:
            employee_name = None

    # Extract project/client identifier
    project = None
    category_match = _PEOPLESOFT_CATEGORY_PATTERN.search(text)
    if category_match:
        project = category_match.group(1).strip()
        logger.debug("Extracted project/category: %s", project)

    entries = []
    lines = text.split("\n")

    # ------------------------------------------------------------------
    # Try MBO Partners format: "timesheet for Aug 04 - Aug 10, 2025"
    # ------------------------------------------------------------------
    mbo_matches = list(_MBO_TITLE_DATE_PATTERN.finditer(text))
    if mbo_matches:
        for mbo_match in mbo_matches:
            start_month, start_day, end_month, end_day, year = mbo_match.groups()
            try:
                start_date = datetime.strptime(f"{start_month} {start_day} {year}", "%b %d %Y").date()
                end_date = datetime.strptime(f"{end_month} {end_day} {year}", "%b %d %Y").date()
            except ValueError:
                continue

            # Search nearby text for "Total XX.XX"
            region_start = max(0, mbo_match.start() - 200)
            region_end = min(len(text), mbo_match.end() + 1000)
            region = text[region_start:region_end]

            hours = None
            total_match = _TOTAL_ONLY_PATTERN.search(region)
            if total_match:
                hours = float(total_match.group(1))

            if hours is None:
                continue

            span = (end_date - start_date).days
            period_type = "weekly" if span <= 7 else "monthly"
            entries.append(TimesheetEntry(
                start_date=start_date, end_date=end_date,
                hours=hours, period_type=period_type, status=None,
            ))

        if entries:
            logger.debug("MBO Partners format: parsed %d entries", len(entries))
            return TimesheetFileData(
                employee_name=employee_name, project=project,
                entries=entries, raw_text=text,
            )

    # ------------------------------------------------------------------
    # Try calendar month view (monthly grids with weekly totals)
    # ------------------------------------------------------------------
    calendar_entries = _parse_calendar_view(text)
    if calendar_entries:
        logger.debug("Calendar view format: parsed %d entries", len(calendar_entries))
        return TimesheetFileData(
            employee_name=employee_name, project=project,
            entries=calendar_entries, raw_text=text,
        )

    # ------------------------------------------------------------------
    # Standard line-by-line date range matching
    # ------------------------------------------------------------------
    for i, line in enumerate(lines):
        line = line.strip()

        # Skip header/filter lines (e.g. "Specific time period" followed by a date range)
        if i > 0 and re.search(r"specific\s+time\s+period|date\s+range|filter", lines[i - 1], re.IGNORECASE):
            continue

        # Try DD-Mon-YYYY format first (primary format)
        match = _DATE_RANGE_PATTERN.search(line)
        if not match:
            # Try "From Sunday MM/DD/YYYY to Saturday MM/DD/YYYY" (PeopleSoft)
            match = _FROM_TO_DATE_PATTERN.search(line)
        if not match:
            # Try parenthesized date range "(MM/DD/YYYY - MM/DD/YYYY)" (Unanet EaZyTyme)
            match = _PAREN_DATE_RANGE.search(line)
        if not match:
            # Try numeric format as fallback
            match = _NUMERIC_DATE_RANGE_PATTERN.search(line)
        if not match:
            continue

        start_date = _parse_date_flexible(match.group(1))
        end_date = _parse_date_flexible(match.group(2))
        if not start_date or not end_date:
            continue

        # Look around for hours: check surrounding lines
        hours = None
        status = None
        # Search this line and a few lines ahead and behind
        search_start = max(0, i - 3)
        search_end = min(i + 6, len(lines))
        search_text = "\n".join(lines[search_start:search_end])

        # For PeopleSoft timesheets, prefer the CWR grid total (excludes CWADJ adjustments)
        cwr_match = _CWR_GRID_TOTAL_PATTERN.search(search_text)
        if cwr_match:
            hours = float(cwr_match.group(1))
        if hours is None or hours == 0:
            # Fallback: sum CWR detail rows ("Approved X.XXX CWR") from a
            # wider region — the "Reported Time Status" table may be far below
            # the date range header.
            region_start = max(0, i - 5)
            region_end = min(len(lines), i + 50)
            detail_region = "\n".join(lines[region_start:region_end])
            detail_matches = _CWR_DETAIL_ROW_PATTERN.findall(detail_region)
            if detail_matches:
                cwr_sum = sum(float(h) for h in detail_matches)
                if cwr_sum > 0:
                    hours = cwr_sum
        if hours is None:
            # Try "Total Hours:" first, then "Reported Hours", then "Totals:", then inline
            hours_match = _HOURS_PATTERN.search(search_text)
            if not hours_match:
                hours_match = _REPORTED_HOURS_PATTERN.search(search_text)
            if not hours_match:
                hours_match = _TOTALS_PATTERN.search(search_text)
            if not hours_match:
                # Try Unanet detail grid row (BILL ... 40.00)
                hours_match = _UNANET_DETAIL_GRID_PATTERN.search(search_text)
            if not hours_match:
                # Try inline hours after date range (Unanet Time-List)
                inline_match = _INLINE_HOURS_AFTER_DATES.search(line)
                if inline_match:
                    hours_match = inline_match
                    # group(3) holds the hours in this pattern
            if hours_match:
                # _INLINE_HOURS_AFTER_DATES uses group(3); all others use group(1)
                if hours_match.re == _INLINE_HOURS_AFTER_DATES:
                    hours = float(hours_match.group(3))
                else:
                    hours = float(hours_match.group(1))

        # Sanity check: reject entries over 160 hours/week or proportional for monthly
        # (catches CWADJ adjustment artifacts and OCR-corrupted totals)
        max_hours = 744 if (end_date - start_date).days > 7 else 160
        if hours is not None and hours > max_hours:
            logger.debug("Rejecting unreasonable hours %.1f for %s - %s", hours, start_date, end_date)
            hours = None

        status_match = _STATUS_PATTERN.search(search_text)
        if status_match:
            status = status_match.group(1).capitalize()

        if hours is None:
            # Skip entries without hours — can't be useful
            logger.debug("Skipping date range %s - %s: no hours found", start_date, end_date)
            continue

        # Determine period type from date span
        span = (end_date - start_date).days
        period_type = "weekly" if span <= 7 else "monthly"

        entries.append(TimesheetEntry(
            start_date=start_date,
            end_date=end_date,
            hours=hours,
            period_type=period_type,
            status=status,
        ))

    logger.debug(
        "Parsed %d entries from text (%d characters), employee: %s, project: %s",
        len(entries), len(text), employee_name, project
    )

    return TimesheetFileData(
        employee_name=employee_name,
        project=project,
        entries=entries,
        raw_text=text,
    )


def _parse_calendar_view(text: str) -> List[TimesheetEntry]:
    """
    Parse a monthly calendar grid view with weekly hour totals.

    Looks for a month+year header (e.g., "June 2025") and weekly totals
    (numbers appearing at line ends, typically 32 or 40). Returns one
    TimesheetEntry per week.
    """
    # Find month + year header (OCR may omit space: "July2025")
    month_match = re.search(
        r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s*(\d{4})\b",
        text, re.IGNORECASE
    )
    if not month_match:
        return []

    month_name = month_match.group(1)
    year = int(month_match.group(2))
    try:
        month_num = datetime.strptime(month_name, "%B").month
    except ValueError:
        return []

    # Pre-clean OCR errors common in calendar grids:
    # - '8' misread as 'S'
    # - '4' misread as '<' (e.g. "40" → "<0")
    cleaned = re.sub(r'\bS\b', '8', text)
    cleaned = re.sub(r'<(\d)', r'4\1', cleaned)

    # Calendar format: rows of daily hours ending with a weekly total, e.g.:
    #   "0 8 8 8 8 0 0 39"
    #   "8 8 8 8 8 0 0 40"
    # OCR may garble some tokens (e.g. ":" instead of "0"), so we extract
    # all numbers from a line and check if the last one is a plausible weekly total.
    weekly_matches = []
    for line in cleaned.split("\n"):
        nums = re.findall(r"\d+(?:\.\d+)?", line)
        # Need at least 8 numbers (7 daily + 1 total) to be a calendar hours row
        if len(nums) < 8:
            continue
        last_num = float(nums[-1])
        # Weekly total should be >= 8 (at least one working day) and <= 80
        # Daily values should all be small (< 24)
        daily_vals = [float(n) for n in nums[:-1]]
        if 8 <= last_num <= 80 and all(v < 24 for v in daily_vals[-7:]):
            weekly_matches.append(str(last_num))

    if not weekly_matches:
        return []

    # Build week boundaries for the month
    _, days_in_month = calendar.monthrange(year, month_num)
    first_day = date(year, month_num, 1)
    last_day = date(year, month_num, days_in_month)

    # Find all Sundays in or overlapping the month
    weeks = []
    # Start from the Sunday on or before the 1st
    week_start = first_day - timedelta(days=(first_day.weekday() + 1) % 7)
    while week_start <= last_day:
        week_end = week_start + timedelta(days=6)
        weeks.append((week_start, week_end))
        week_start += timedelta(days=7)

    entries = []
    for idx, hours_str in enumerate(weekly_matches):
        if idx >= len(weeks):
            break
        hours = float(hours_str)
        if hours <= 0:
            continue
        ws, we = weeks[idx]
        entries.append(TimesheetEntry(
            start_date=ws, end_date=we,
            hours=hours, period_type="weekly", status=None,
        ))

    return entries


def normalize_to_week_start(d: date) -> date:
    """
    Normalize a date to its containing week's Sunday.

    The Excel weekly rows use Sunday as the week start. This function
    finds the Sunday on or before the given date.

    Args:
        d: Any date.

    Returns:
        The Sunday on or before d.
    """
    # weekday(): Mon=0 ... Sun=6
    days_since_sunday = (d.weekday() + 1) % 7
    return d - timedelta(days=days_since_sunday)
